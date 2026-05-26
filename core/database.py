"""
core/database.py - Gestión de base de datos SQLite
"""
import sqlite3
import os
import shutil
import datetime


class Database:
    def __init__(self, db_path):
        self.db_path = db_path
        self.conn = None

    def connect(self):
        self.conn = sqlite3.connect(self.db_path)
        self.conn.row_factory = sqlite3.Row
        self.conn.execute("PRAGMA foreign_keys = ON")
        return self.conn

    def get_conn(self):
        if not self.conn:
            self.connect()
        return self.conn

    def initialize(self):
        conn = self.connect()
        cur = conn.cursor()

        cur.executescript("""
        CREATE TABLE IF NOT EXISTS usuarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            nombre TEXT,
            rol TEXT DEFAULT 'operador',
            activo INTEGER DEFAULT 1,
            creado TEXT DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS inventario (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fabricante TEXT NOT NULL,
            referencia TEXT NOT NULL,
            descripcion TEXT,
            cantidad REAL DEFAULT 0,
            valor_unitario REAL DEFAULT 0,
            valor_total REAL DEFAULT 0,
            stock_minimo REAL DEFAULT 2,
            stock_critico REAL DEFAULT 1,
            ubicacion TEXT,
            maquina TEXT,
            imagen_path TEXT,
            observaciones TEXT,
            activo INTEGER DEFAULT 1,
            creado TEXT DEFAULT CURRENT_TIMESTAMP,
            actualizado TEXT DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS entradas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            inventario_id INTEGER REFERENCES inventario(id),
            fabricante TEXT,
            referencia TEXT,
            descripcion TEXT,
            cantidad REAL,
            valor_unitario REAL,
            valor_total REAL,
            ubicacion TEXT,
            maquina TEXT,
            observaciones TEXT,
            usuario TEXT,
            fecha TEXT DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS salidas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            inventario_id INTEGER REFERENCES inventario(id),
            fabricante TEXT,
            referencia TEXT,
            cantidad REAL,
            valor_unitario REAL,
            costo_consumido REAL,
            maquina TEXT,
            area TEXT,
            tecnico TEXT,
            motivo TEXT,
            observaciones TEXT,
            usuario TEXT,
            fecha TEXT DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS alertas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            inventario_id INTEGER REFERENCES inventario(id),
            tipo TEXT,
            mensaje TEXT,
            leida INTEGER DEFAULT 0,
            fecha TEXT DEFAULT CURRENT_TIMESTAMP
        );
        """)

        # Usuario admin por defecto
        cur.execute("SELECT id FROM usuarios WHERE username='admin'")
        if not cur.fetchone():
            cur.execute(
                "INSERT INTO usuarios (username, password, nombre, rol) VALUES (?,?,?,?)",
                ("admin", "admin123", "Administrador", "admin")
            )

        conn.commit()

    def migrate_excel(self, excel_path):
        """Migrar datos del Excel al inventario si no han sido migrados."""
        if not os.path.exists(excel_path):
            return
        conn = self.get_conn()
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM inventario")
        if cur.fetchone()[0] > 0:
            return  # Ya migrado

        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            if "INVENTARIO" in wb.sheetnames:
                ws = wb["INVENTARIO"]
                rows = list(ws.iter_rows(values_only=True))
                # Saltar encabezado
                for row in rows[1:]:
                    if not row or not row[0]:
                        continue
                    fabricante = str(row[0]).strip() if row[0] else ""
                    referencia = str(row[1]).strip() if row[1] else ""
                    cantidad = float(row[2]) if row[2] else 0
                    valor_unitario = float(row[3]) if row[3] else 0
                    valor_total = cantidad * valor_unitario

                    if fabricante and referencia:
                        cur.execute("""
                            INSERT INTO inventario (fabricante, referencia, cantidad, valor_unitario, valor_total)
                            VALUES (?,?,?,?,?)
                        """, (fabricante, referencia, cantidad, valor_unitario, valor_total))
            conn.commit()
        except Exception as e:
            print(f"Advertencia migración Excel: {e}")

    # ─── INVENTARIO ────────────────────────────────────────────────────────────

    def get_inventario(self, search="", filtro_estado=""):
        conn = self.get_conn()
        query = "SELECT * FROM inventario WHERE activo=1"
        params = []
        if search:
            query += " AND (fabricante LIKE ? OR referencia LIKE ? OR descripcion LIKE ? OR maquina LIKE ?)"
            s = f"%{search}%"
            params += [s, s, s, s]
        query += " ORDER BY fabricante, referencia"
        return conn.execute(query, params).fetchall()

    def get_item_by_id(self, item_id):
        return self.get_conn().execute("SELECT * FROM inventario WHERE id=?", (item_id,)).fetchone()

    def item_exists(self, fabricante, referencia):
        row = self.get_conn().execute(
            "SELECT id FROM inventario WHERE TRIM(LOWER(fabricante))=TRIM(LOWER(?)) AND TRIM(LOWER(referencia))=TRIM(LOWER(?))",
            (fabricante, referencia)
        ).fetchone()
        return row["id"] if row else None

    def add_item(self, data):
        conn = self.get_conn()
        conn.execute("""
            INSERT INTO inventario (fabricante,referencia,descripcion,cantidad,valor_unitario,
            valor_total,stock_minimo,stock_critico,ubicacion,maquina,imagen_path,observaciones)
            VALUES (:fabricante,:referencia,:descripcion,:cantidad,:valor_unitario,
            :valor_total,:stock_minimo,:stock_critico,:ubicacion,:maquina,:imagen_path,:observaciones)
        """, data)
        conn.commit()

    def update_stock(self, item_id, nueva_cantidad, nuevo_valor_unitario=None):
        conn = self.get_conn()
        if nuevo_valor_unitario is not None:
            conn.execute("""
                UPDATE inventario SET cantidad=?, valor_unitario=?,
                valor_total=cantidad*?, actualizado=CURRENT_TIMESTAMP WHERE id=?
            """, (nueva_cantidad, nuevo_valor_unitario, nuevo_valor_unitario, item_id))
        else:
            conn.execute("""
                UPDATE inventario SET cantidad=?,
                valor_total=cantidad*valor_unitario, actualizado=CURRENT_TIMESTAMP WHERE id=?
            """, (nueva_cantidad, item_id))
        conn.commit()

    def update_item(self, item_id, data):
        conn = self.get_conn()
        conn.execute("""
            UPDATE inventario SET fabricante=:fabricante, referencia=:referencia,
            descripcion=:descripcion, valor_unitario=:valor_unitario,
            valor_total=cantidad*:valor_unitario, stock_minimo=:stock_minimo,
            stock_critico=:stock_critico, ubicacion=:ubicacion, maquina=:maquina,
            imagen_path=:imagen_path, observaciones=:observaciones,
            actualizado=CURRENT_TIMESTAMP WHERE id=:id
        """, {**data, "id": item_id})
        conn.commit()

    # ─── ENTRADAS ──────────────────────────────────────────────────────────────

    def registrar_entrada(self, data, usuario):
        conn = self.get_conn()
        item_id = self.item_exists(data["fabricante"], data["referencia"])

        if item_id:
            row = conn.execute("SELECT cantidad, valor_unitario FROM inventario WHERE id=?", (item_id,)).fetchone()
            nueva_cant = row["cantidad"] + data["cantidad"]
            self.update_stock(item_id, nueva_cant, data.get("valor_unitario") or row["valor_unitario"])
        else:
            item_data = {
                "fabricante": data["fabricante"],
                "referencia": data["referencia"],
                "descripcion": data.get("descripcion", ""),
                "cantidad": data["cantidad"],
                "valor_unitario": data.get("valor_unitario", 0),
                "valor_total": data["cantidad"] * data.get("valor_unitario", 0),
                "stock_minimo": data.get("stock_minimo", 2),
                "stock_critico": data.get("stock_critico", 1),
                "ubicacion": data.get("ubicacion", ""),
                "maquina": data.get("maquina", ""),
                "imagen_path": data.get("imagen_path", ""),
                "observaciones": data.get("observaciones", ""),
            }
            self.add_item(item_data)
            item_id = self.item_exists(data["fabricante"], data["referencia"])

        conn.execute("""
            INSERT INTO entradas (inventario_id,fabricante,referencia,descripcion,cantidad,
            valor_unitario,valor_total,ubicacion,maquina,observaciones,usuario,fecha)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
        """, (item_id, data["fabricante"], data["referencia"], data.get("descripcion",""),
              data["cantidad"], data.get("valor_unitario",0),
              data["cantidad"]*data.get("valor_unitario",0),
              data.get("ubicacion",""), data.get("maquina",""),
              data.get("observaciones",""), usuario, data.get("fecha", datetime.datetime.now().isoformat())))
        conn.commit()
        self._check_alertas(item_id)

    # ─── SALIDAS ───────────────────────────────────────────────────────────────

    def registrar_salida(self, data, usuario):
        conn = self.get_conn()
        item_id = self.item_exists(data["fabricante"], data["referencia"])
        if not item_id:
            raise ValueError(f"Producto {data['referencia']} no encontrado en inventario.")

        row = conn.execute("SELECT cantidad, valor_unitario FROM inventario WHERE id=?", (item_id,)).fetchone()
        if row["cantidad"] < data["cantidad"]:
            raise ValueError(f"Stock insuficiente. Disponible: {row['cantidad']}")

        nueva_cant = row["cantidad"] - data["cantidad"]
        self.update_stock(item_id, nueva_cant)

        conn.execute("""
            INSERT INTO salidas (inventario_id,fabricante,referencia,cantidad,valor_unitario,
            costo_consumido,maquina,area,tecnico,motivo,observaciones,usuario,fecha)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (item_id, data["fabricante"], data["referencia"], data["cantidad"],
              row["valor_unitario"], data["cantidad"]*row["valor_unitario"],
              data.get("maquina",""), data.get("area",""), data.get("tecnico",""),
              data.get("motivo",""), data.get("observaciones",""),
              usuario, data.get("fecha", datetime.datetime.now().isoformat())))
        conn.commit()
        self._check_alertas(item_id)

    # ─── ALERTAS ───────────────────────────────────────────────────────────────

    def _check_alertas(self, item_id):
        conn = self.get_conn()
        row = conn.execute("SELECT * FROM inventario WHERE id=?", (item_id,)).fetchone()
        if not row:
            return
        if row["cantidad"] == 0:
            self._add_alerta(item_id, "AGOTADO", f"AGOTADO: {row['fabricante']} {row['referencia']}")
        elif row["cantidad"] <= row["stock_critico"]:
            self._add_alerta(item_id, "CRITICO", f"Stock CRÍTICO: {row['fabricante']} {row['referencia']} ({row['cantidad']} uds)")
        elif row["cantidad"] <= row["stock_minimo"]:
            self._add_alerta(item_id, "BAJO", f"Stock BAJO: {row['fabricante']} {row['referencia']} ({row['cantidad']} uds)")

    def _add_alerta(self, item_id, tipo, mensaje):
        conn = self.get_conn()
        conn.execute("INSERT INTO alertas (inventario_id, tipo, mensaje) VALUES (?,?,?)",
                     (item_id, tipo, mensaje))
        conn.commit()

    def get_alertas(self, solo_no_leidas=True):
        q = "SELECT a.*, i.fabricante, i.referencia FROM alertas a LEFT JOIN inventario i ON a.inventario_id=i.id"
        if solo_no_leidas:
            q += " WHERE a.leida=0"
        q += " ORDER BY a.fecha DESC LIMIT 50"
        return self.get_conn().execute(q).fetchall()

    def marcar_alertas_leidas(self):
        self.get_conn().execute("UPDATE alertas SET leida=1")
        self.get_conn().commit()

    # ─── REPORTES ──────────────────────────────────────────────────────────────

    def get_entradas(self, desde=None, hasta=None):
        q = "SELECT * FROM entradas WHERE 1=1"
        p = []
        if desde:
            q += " AND fecha >= ?"
            p.append(desde)
        if hasta:
            q += " AND fecha <= ?"
            p.append(hasta)
        q += " ORDER BY fecha DESC"
        return self.get_conn().execute(q, p).fetchall()

    def get_salidas(self, desde=None, hasta=None):
        q = "SELECT * FROM salidas WHERE 1=1"
        p = []
        if desde:
            q += " AND fecha >= ?"
            p.append(desde)
        if hasta:
            q += " AND fecha <= ?"
            p.append(hasta)
        q += " ORDER BY fecha DESC"
        return self.get_conn().execute(q, p).fetchall()

    def get_stats(self):
        conn = self.get_conn()
        total_items = conn.execute("SELECT COUNT(*) FROM inventario WHERE activo=1").fetchone()[0]
        valor_total = conn.execute("SELECT SUM(valor_total) FROM inventario WHERE activo=1").fetchone()[0] or 0
        agotados = conn.execute("SELECT COUNT(*) FROM inventario WHERE activo=1 AND cantidad=0").fetchone()[0]
        criticos = conn.execute("SELECT COUNT(*) FROM inventario WHERE activo=1 AND cantidad>0 AND cantidad<=stock_critico").fetchone()[0]
        bajos = conn.execute("SELECT COUNT(*) FROM inventario WHERE activo=1 AND cantidad>stock_critico AND cantidad<=stock_minimo").fetchone()[0]
        alertas_pend = conn.execute("SELECT COUNT(*) FROM alertas WHERE leida=0").fetchone()[0]
        return {
            "total_items": total_items,
            "valor_total": valor_total,
            "agotados": agotados,
            "criticos": criticos,
            "bajos": bajos,
            "alertas_pendientes": alertas_pend,
        }

    def get_consumo_por_maquina(self):
        return self.get_conn().execute("""
            SELECT maquina, SUM(costo_consumido) as total_costo, COUNT(*) as n_movimientos
            FROM salidas WHERE maquina != '' GROUP BY maquina ORDER BY total_costo DESC
        """).fetchall()

    def get_productos_mas_usados(self, limit=10):
        return self.get_conn().execute("""
            SELECT fabricante, referencia, SUM(cantidad) as total_salidas, SUM(costo_consumido) as costo_total
            FROM salidas GROUP BY fabricante, referencia ORDER BY total_salidas DESC LIMIT ?
        """, (limit,)).fetchall()

    # ─── BACKUPS ───────────────────────────────────────────────────────────────

    def hacer_backup(self, backup_dir):
        os.makedirs(backup_dir, exist_ok=True)
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        dest = os.path.join(backup_dir, f"inventario_backup_{ts}.db")
        shutil.copy2(self.db_path, dest)
        return dest

    # ─── USUARIOS ──────────────────────────────────────────────────────────────

    def login(self, username, password):
        row = self.get_conn().execute(
            "SELECT * FROM usuarios WHERE username=? AND password=? AND activo=1",
            (username, password)
        ).fetchone()
        return dict(row) if row else None
