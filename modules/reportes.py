"""
modules/reportes.py - Módulo de reportes y analíticas
"""
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import datetime, os

BG_DARK  = "#0D1117"
BG_CARD  = "#161B22"
BG_INPUT = "#21262D"
ACCENT   = "#00D4FF"
TEXT     = "#E6EDF3"
TEXT_DIM = "#8B949E"
SUCCESS  = "#3FB950"
DANGER   = "#FF4C4C"
WARNING  = "#D29922"
BORDER   = "#30363D"


class ReportesModule:
    def __init__(self, parent, db):
        self.parent = parent
        self.db = db
        self._build()

    def _build(self):
        # Filtros de fechas
        top = tk.Frame(self.parent, bg=BG_DARK)
        top.pack(fill="x", pady=(0, 12))

        tk.Label(top, text="Desde:", font=("Arial", 9), bg=BG_DARK, fg=TEXT_DIM).pack(side="left")
        self.entry_desde = tk.Entry(top, font=("Arial", 10), bg=BG_CARD, fg=TEXT,
                                    insertbackground=TEXT, relief="flat", bd=6, width=14)
        self.entry_desde.pack(side="left", padx=4, ipady=4)
        mes_inicio = datetime.datetime.now().replace(day=1).strftime("%Y-%m-%d")
        self.entry_desde.insert(0, mes_inicio)

        tk.Label(top, text="Hasta:", font=("Arial", 9), bg=BG_DARK, fg=TEXT_DIM).pack(side="left", padx=(10, 0))
        self.entry_hasta = tk.Entry(top, font=("Arial", 10), bg=BG_CARD, fg=TEXT,
                                    insertbackground=TEXT, relief="flat", bd=6, width=14)
        self.entry_hasta.pack(side="left", padx=4, ipady=4)
        self.entry_hasta.insert(0, datetime.datetime.now().strftime("%Y-%m-%d"))

        tk.Button(top, text="🔄 Generar", font=("Arial", 9), bg=ACCENT, fg=BG_DARK,
                  relief="flat", cursor="hand2", padx=10, pady=4,
                  command=self._refresh).pack(side="left", padx=8)

        # Export buttons
        tk.Button(top, text="📄 PDF Completo", font=("Arial", 9), bg=BORDER, fg=TEXT,
                  relief="flat", cursor="hand2", padx=10, pady=4,
                  command=self._export_pdf).pack(side="right", padx=4)
        tk.Button(top, text="📊 Excel Completo", font=("Arial", 9), bg=BORDER, fg=TEXT,
                  relief="flat", cursor="hand2", padx=10, pady=4,
                  command=self._export_excel).pack(side="right", padx=4)

        # Tabs
        nb = ttk.Notebook(self.parent)
        style = ttk.Style()
        style.configure("TNotebook", background=BG_DARK, borderwidth=0)
        style.configure("TNotebook.Tab", background=BG_CARD, foreground=TEXT_DIM,
                         padding=(12, 6), font=("Arial", 9))
        style.map("TNotebook.Tab", background=[("selected", "#1F6FEB")],
                  foreground=[("selected", TEXT)])
        nb.pack(fill="both", expand=True)

        self.tab_resumen   = tk.Frame(nb, bg=BG_DARK)
        self.tab_consumo   = tk.Frame(nb, bg=BG_DARK)
        self.tab_entradas  = tk.Frame(nb, bg=BG_DARK)
        self.tab_salidas   = tk.Frame(nb, bg=BG_DARK)
        self.tab_criticos  = tk.Frame(nb, bg=BG_DARK)

        nb.add(self.tab_resumen,  text="📊 Resumen")
        nb.add(self.tab_consumo,  text="🔧 Consumo x Máquina")
        nb.add(self.tab_entradas, text="📥 Entradas")
        nb.add(self.tab_salidas,  text="📤 Salidas")
        nb.add(self.tab_criticos, text="🚨 Críticos")

        self._refresh()

    def _get_dates(self):
        desde = self.entry_desde.get().strip() or None
        hasta = self.entry_hasta.get().strip() or None
        return desde, hasta

    def _refresh(self):
        desde, hasta = self._get_dates()
        self._build_resumen()
        self._build_consumo()
        self._build_entradas(desde, hasta)
        self._build_salidas(desde, hasta)
        self._build_criticos()

    def _build_resumen(self):
        for w in self.tab_resumen.winfo_children():
            w.destroy()

        stats = self.db.get_stats()
        salidas = self.db.get_salidas()
        entradas = self.db.get_entradas()

        total_consumido = sum(r["costo_consumido"] for r in salidas)
        total_ingresado = sum(r["valor_total"] for r in entradas)

        cards = tk.Frame(self.tab_resumen, bg=BG_DARK)
        cards.pack(fill="x", pady=8, padx=8)

        data_cards = [
            ("Total Productos", stats["total_items"], ACCENT, "📦"),
            ("Valor en Stock", f"${stats['valor_total']:,.0f}", SUCCESS, "💰"),
            ("Total Consumido", f"${total_consumido:,.0f}", WARNING, "📤"),
            ("Total Ingresado", f"${total_ingresado:,.0f}", "#8B5CF6", "📥"),
        ]
        for title, val, color, icon in data_cards:
            f = tk.Frame(cards, bg=BG_CARD, padx=18, pady=14)
            f.pack(side="left", fill="both", expand=True, padx=4)
            tk.Label(f, text=icon, font=("Arial", 22), bg=BG_CARD, fg=color).pack()
            tk.Label(f, text=str(val), font=("Arial", 20, "bold"), bg=BG_CARD, fg=color).pack(pady=4)
            tk.Label(f, text=title, font=("Arial", 9), bg=BG_CARD, fg=TEXT_DIM).pack()
            tk.Frame(f, bg=color, height=3).pack(fill="x", side="bottom", pady=(8,0))

        # Top productos más consumidos
        frame = tk.Frame(self.tab_resumen, bg=BG_CARD, padx=8, pady=8)
        frame.pack(fill="both", expand=True, padx=8, pady=4)

        tk.Label(frame, text="🔝 TOP 10 PRODUCTOS MÁS CONSUMIDOS", font=("Arial", 10, "bold"),
                 bg=BG_CARD, fg=ACCENT).pack(anchor="w", pady=(0,8))

        cols = ("Fabricante", "Referencia", "Unidades Consumidas", "Costo Total")
        tree = ttk.Treeview(frame, columns=cols, show="headings", height=10)
        style = ttk.Style()
        style.configure("Treeview", background=BG_CARD, foreground=TEXT,
                         fieldbackground=BG_CARD, rowheight=26)
        style.configure("Treeview.Heading", background="#21262D", foreground=TEXT_DIM,
                         font=("Arial", 9, "bold"))
        style.map("Treeview", background=[("selected", "#1F6FEB")])

        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=160)

        for row in self.db.get_productos_mas_usados():
            tree.insert("", "end", values=(
                row["fabricante"], row["referencia"],
                row["total_salidas"], f"${row['costo_total']:,.0f}"
            ))
        tree.pack(fill="both", expand=True)

    def _build_consumo(self):
        for w in self.tab_consumo.winfo_children():
            w.destroy()

        frame = tk.Frame(self.tab_consumo, bg=BG_CARD, padx=8, pady=8)
        frame.pack(fill="both", expand=True, padx=8, pady=8)

        tk.Label(frame, text="🔧 CONSUMO POR MÁQUINA / ÁREA", font=("Arial", 10, "bold"),
                 bg=BG_CARD, fg=ACCENT).pack(anchor="w", pady=(0,8))

        cols = ("Máquina / Área", "Costo Total Consumido", "N° Movimientos")
        tree = ttk.Treeview(frame, columns=cols, show="headings", height=16)
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=200)

        rows = self.db.get_consumo_por_maquina()
        if not rows:
            tk.Label(frame, text="Sin datos disponibles.\nRegistre salidas con máquina asociada.",
                     font=("Arial", 11), bg=BG_CARD, fg=TEXT_DIM).pack(expand=True)
            return

        for r in rows:
            tree.insert("", "end", values=(r["maquina"], f"${r['total_costo']:,.0f}", r["n_movimientos"]))
        tree.pack(fill="both", expand=True)

    def _build_table_tab(self, tab, title, cols, rows_data, color=ACCENT):
        for w in tab.winfo_children():
            w.destroy()

        frame = tk.Frame(tab, bg=BG_CARD, padx=8, pady=8)
        frame.pack(fill="both", expand=True, padx=8, pady=8)

        tk.Label(frame, text=title, font=("Arial", 10, "bold"),
                 bg=BG_CARD, fg=color).pack(anchor="w", pady=(0,8))

        tree = ttk.Treeview(frame, columns=cols, show="headings", height=18)
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=max(120, 800//len(cols)))

        for row in rows_data:
            tree.insert("", "end", values=row)

        sb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=sb.set)
        tree.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        if not rows_data:
            tk.Label(frame, text="Sin datos en el período seleccionado.",
                     font=("Arial", 10), bg=BG_CARD, fg=TEXT_DIM).pack()

    def _build_entradas(self, desde, hasta):
        rows = self.db.get_entradas(desde, hasta)
        data = [(r["fecha"][:16], r["fabricante"], r["referencia"],
                 r["cantidad"], f"${r['valor_unitario']:,.0f}",
                 f"${r['valor_total']:,.0f}", r["ubicacion"], r["maquina"], r["usuario"]) for r in rows]
        cols = ("Fecha", "Fabricante", "Referencia", "Cantidad", "Valor Unit.", "Total", "Ubicación", "Máquina", "Usuario")
        self._build_table_tab(self.tab_entradas, "📥 HISTORIAL DE ENTRADAS", cols, data, SUCCESS)

    def _build_salidas(self, desde, hasta):
        rows = self.db.get_salidas(desde, hasta)
        data = [(r["fecha"][:16], r["fabricante"], r["referencia"],
                 r["cantidad"], f"${r['costo_consumido']:,.0f}",
                 r["maquina"], r["area"], r["tecnico"], r["motivo"], r["usuario"]) for r in rows]
        cols = ("Fecha", "Fabricante", "Referencia", "Cantidad", "Costo", "Máquina", "Área", "Técnico", "Motivo", "Usuario")
        self._build_table_tab(self.tab_salidas, "📤 HISTORIAL DE SALIDAS", cols, data, WARNING)

    def _build_criticos(self):
        items = self.db.get_inventario()
        data = []
        for row in items:
            qty = row["cantidad"]
            if qty == 0: estado = "AGOTADO"
            elif qty <= row["stock_critico"]: estado = "CRÍTICO"
            elif qty <= row["stock_minimo"]: estado = "BAJO"
            else: continue
            data.append((row["fabricante"], row["referencia"], qty,
                         row["stock_minimo"], row["stock_critico"],
                         f"${row['valor_unitario']:,.0f}", estado))
        cols = ("Fabricante", "Referencia", "Stock Actual", "Stock Mínimo", "Stock Crítico", "Valor Unit.", "Estado")
        self._build_table_tab(self.tab_criticos, "🚨 PRODUCTOS CRÍTICOS Y AGOTADOS", cols, data, DANGER)

    def _export_pdf(self):
        try:
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        except ImportError:
            messagebox.showerror("Error", "reportlab no instalado.\nInstale con: pip install reportlab")
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")],
            initialfile=f"inventario_reporte_{datetime.datetime.now().strftime('%Y%m%d')}.pdf"
        )
        if not path:
            return

        doc = SimpleDocTemplate(path, pagesize=landscape(letter))
        styles = getSampleStyleSheet()
        elements = []

        # Título
        title_style = ParagraphStyle("title", fontSize=16, textColor=colors.HexColor("#00D4FF"),
                                      spaceAfter=12, fontName="Helvetica-Bold")
        elements.append(Paragraph("REPORTE DE INVENTARIO INDUSTRIAL", title_style))
        elements.append(Paragraph(f"Generado: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}",
                                   styles["Normal"]))
        elements.append(Spacer(1, 12))

        # Stats
        stats = self.db.get_stats()
        stats_data = [
            ["Total Productos", "Valor en Stock", "Agotados", "Críticos", "Bajos"],
            [str(stats["total_items"]), f"${stats['valor_total']:,.0f}",
             str(stats["agotados"]), str(stats["criticos"]), str(stats["bajos"])]
        ]
        t = Table(stats_data, colWidths=[120]*5)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#161B22")),
            ("TEXTCOLOR", (0,0), (-1,0), colors.HexColor("#00D4FF")),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
            ("ALIGN", (0,0), (-1,-1), "CENTER"),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F5F5F5")]),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 16))

        # Inventario completo
        elements.append(Paragraph("INVENTARIO GENERAL", ParagraphStyle("h2", fontSize=12,
                                    fontName="Helvetica-Bold", spaceAfter=8)))
        items = self.db.get_inventario()
        if items:
            data = [["Fabricante", "Referencia", "Descripción", "Cantidad", "Val. Unit.", "Val. Total", "Estado"]]
            for row in items:
                qty = row["cantidad"]
                if qty == 0: estado = "AGOTADO"
                elif qty <= row["stock_critico"]: estado = "CRÍTICO"
                elif qty <= row["stock_minimo"]: estado = "BAJO"
                else: estado = "Normal"
                data.append([row["fabricante"], row["referencia"],
                              (row["descripcion"] or "")[:30],
                              qty, f"${row['valor_unitario']:,.0f}",
                              f"${row['valor_total']:,.0f}", estado])

            t2 = Table(data, colWidths=[100, 100, 160, 60, 70, 80, 70])
            t2.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#0D1117")),
                ("TEXTCOLOR", (0,0), (-1,0), colors.white),
                ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE", (0,0), (-1,-1), 8),
                ("GRID", (0,0), (-1,-1), 0.3, colors.grey),
                ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F9F9F9")]),
            ]))
            elements.append(t2)

        doc.build(elements)
        messagebox.showinfo("PDF Generado", f"Reporte guardado en:\n{path}")

    def _export_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showerror("Error", "openpyxl no instalado.")
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"reporte_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx"
        )
        if not path:
            return

        wb = openpyxl.Workbook()
        desde, hasta = self._get_dates()

        # Hoja Inventario
        ws = wb.active
        ws.title = "Inventario"
        h_fill = PatternFill("solid", start_color="0D1117")
        h_font = Font(bold=True, color="00D4FF")

        hdrs = ["Fabricante", "Referencia", "Descripción", "Cantidad", "Val. Unit.", "Val. Total",
                "Ubicación", "Máquina", "Stock Mín.", "Stock Crít.", "Estado"]
        for col, h in enumerate(hdrs, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = h_font
            cell.fill = h_fill

        for row in self.db.get_inventario():
            qty = row["cantidad"]
            if qty == 0: estado = "AGOTADO"
            elif qty <= row["stock_critico"]: estado = "CRÍTICO"
            elif qty <= row["stock_minimo"]: estado = "BAJO"
            else: estado = "Normal"
            ws.append([row["fabricante"], row["referencia"], row["descripcion"],
                       qty, row["valor_unitario"], row["valor_total"],
                       row["ubicacion"], row["maquina"],
                       row["stock_minimo"], row["stock_critico"], estado])

        # Hoja Entradas
        ws2 = wb.create_sheet("Entradas")
        hdrs2 = ["Fecha", "Fabricante", "Referencia", "Cantidad", "Valor Unit.", "Total", "Máquina", "Usuario"]
        for col, h in enumerate(hdrs2, 1):
            ws2.cell(row=1, column=col, value=h).font = h_font
        for r in self.db.get_entradas(desde, hasta):
            ws2.append([r["fecha"][:16], r["fabricante"], r["referencia"],
                        r["cantidad"], r["valor_unitario"], r["valor_total"],
                        r["maquina"], r["usuario"]])

        # Hoja Salidas
        ws3 = wb.create_sheet("Salidas")
        hdrs3 = ["Fecha", "Fabricante", "Referencia", "Cantidad", "Costo", "Máquina", "Técnico", "Motivo", "Usuario"]
        for col, h in enumerate(hdrs3, 1):
            ws3.cell(row=1, column=col, value=h).font = h_font
        for r in self.db.get_salidas(desde, hasta):
            ws3.append([r["fecha"][:16], r["fabricante"], r["referencia"],
                        r["cantidad"], r["costo_consumido"],
                        r["maquina"], r["tecnico"], r["motivo"], r["usuario"]])

        for ws_ in [ws, ws2, ws3]:
            for col in ws_.columns:
                mx = max(len(str(c.value or "")) for c in col)
                ws_.column_dimensions[col[0].column_letter].width = min(mx + 4, 40)

        wb.save(path)
        messagebox.showinfo("Excel Generado", f"Reporte guardado en:\n{path}")
