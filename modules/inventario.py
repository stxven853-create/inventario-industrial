"""
modules/inventario.py - Panel de inventario general
"""
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import os

BG_DARK  = "#0D1117"
BG_CARD  = "#161B22"
BG_INPUT = "#21262D"
ACCENT   = "#00D4FF"
TEXT     = "#E6EDF3"
TEXT_DIM = "#8B949E"
SUCCESS  = "#3FB950"
DANGER   = "#FF4C4C"
WARNING  = "#D29922"
ORANGE   = "#FFA500"
BORDER   = "#30363D"


class InventarioModule:
    def __init__(self, parent, db, user):
        self.parent = parent
        self.db = db
        self.user = user
        self.selected_id = None
        self._build()

    def _build(self):
        # ─── Barra superior: búsqueda y filtros ─────────────────────────
        top = tk.Frame(self.parent, bg=BG_DARK)
        top.pack(fill="x", pady=(0, 10))

        tk.Label(top, text="🔍", font=("Arial", 12), bg=BG_DARK, fg=TEXT_DIM).pack(side="left")
        self.entry_search = tk.Entry(top, font=("Arial", 10), bg=BG_CARD, fg=TEXT,
                                     insertbackground=TEXT, relief="flat", bd=6, width=30)
        self.entry_search.pack(side="left", padx=6, ipady=5)
        self.entry_search.bind("<KeyRelease>", lambda e: self._load_table())
        self.entry_search.insert(0, "")

        # Filtro estado
        tk.Label(top, text="Estado:", font=("Arial", 9), bg=BG_DARK, fg=TEXT_DIM).pack(side="left", padx=(12,4))
        self.filter_var = tk.StringVar(value="Todos")
        filter_cb = ttk.Combobox(top, textvariable=self.filter_var, width=12,
                                  values=["Todos", "Normal", "Bajo", "Crítico", "Agotado"],
                                  state="readonly")
        filter_cb.pack(side="left")
        filter_cb.bind("<<ComboboxSelected>>", lambda e: self._load_table())

        # Botones
        tk.Button(top, text="➕ Nueva Entrada", font=("Arial", 9), bg=SUCCESS, fg=BG_DARK,
                  relief="flat", cursor="hand2", padx=10, pady=4,
                  command=self._quick_add).pack(side="right", padx=4)
        tk.Button(top, text="📥 Exportar Excel", font=("Arial", 9), bg=BORDER, fg=TEXT,
                  relief="flat", cursor="hand2", padx=10, pady=4,
                  command=self._export_excel).pack(side="right", padx=4)
        tk.Button(top, text="🔄 Actualizar", font=("Arial", 9), bg=BORDER, fg=TEXT_DIM,
                  relief="flat", cursor="hand2", padx=10, pady=4,
                  command=self._load_table).pack(side="right", padx=4)

        # ─── Layout: tabla + panel detalle ──────────────────────────────
        content = tk.Frame(self.parent, bg=BG_DARK)
        content.pack(fill="both", expand=True)

        # Tabla
        table_frame = tk.Frame(content, bg=BG_CARD, padx=4, pady=4)
        table_frame.pack(side="left", fill="both", expand=True)

        cols = ("ID", "Fabricante", "Referencia", "Descripción", "Cantidad",
                "Val. Unit.", "Val. Total", "Ubicación", "Máquina", "Estado")
        self.tree = ttk.Treeview(table_frame, columns=cols, show="headings", height=24)

        style = ttk.Style()
        style.configure("Treeview", background=BG_CARD, foreground=TEXT,
                         fieldbackground=BG_CARD, rowheight=26, borderwidth=0)
        style.configure("Treeview.Heading", background="#21262D", foreground=TEXT_DIM,
                         font=("Arial", 9, "bold"))
        style.map("Treeview", background=[("selected", "#1F6FEB")])

        self.tree.tag_configure("normal", foreground=TEXT)
        self.tree.tag_configure("bajo", foreground=ORANGE)
        self.tree.tag_configure("critico", foreground=WARNING)
        self.tree.tag_configure("agotado", foreground=DANGER)

        widths = [40, 100, 100, 160, 70, 80, 90, 90, 100, 80]
        for c, w in zip(cols, widths):
            self.tree.heading(c, text=c, command=lambda _c=c: self._sort(_c))
            self.tree.column(c, width=w, minwidth=40)

        sb_y = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        sb_x = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=sb_y.set, xscrollcommand=sb_x.set)
        sb_y.pack(side="right", fill="y")
        sb_x.pack(side="bottom", fill="x")
        self.tree.pack(fill="both", expand=True)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.tree.bind("<Double-1>", self._on_double_click)

        # Panel detalle lateral
        self.detail = tk.Frame(content, bg=BG_CARD, width=220, padx=12, pady=12)
        self.detail.pack(side="right", fill="y", padx=(8, 0))
        self.detail.pack_propagate(False)

        tk.Label(self.detail, text="DETALLE DEL PRODUCTO", font=("Arial", 9, "bold"),
                 bg=BG_CARD, fg=ACCENT).pack(anchor="w")

        self.img_label = tk.Label(self.detail, text="📦", font=("Arial", 40),
                                   bg=BG_CARD, fg=TEXT_DIM)
        self.img_label.pack(pady=12)

        self.detail_text = tk.Text(self.detail, font=("Arial", 9), bg=BG_CARD, fg=TEXT,
                                    relief="flat", height=18, wrap="word", state="disabled")
        self.detail_text.pack(fill="both", expand=True)

        # Contador
        self.lbl_count = tk.Label(self.parent, text="", font=("Arial", 8),
                                   bg=BG_DARK, fg=TEXT_DIM)
        self.lbl_count.pack(side="bottom", anchor="w", pady=4)

        self._load_table()

    def _load_table(self):
        self.tree.delete(*self.tree.get_children())
        search = self.entry_search.get().strip()
        items = self.db.get_inventario(search=search)
        estado_filter = self.filter_var.get()

        count = 0
        for row in items:
            qty = row["cantidad"]
            sc = row["stock_critico"]
            sm = row["stock_minimo"]

            if qty == 0:
                tag, estado = "agotado", "AGOTADO"
            elif qty <= sc:
                tag, estado = "critico", "CRÍTICO"
            elif qty <= sm:
                tag, estado = "bajo", "BAJO"
            else:
                tag, estado = "normal", "Normal"

            if estado_filter != "Todos":
                map_f = {"Normal": "normal", "Bajo": "bajo", "Crítico": "critico", "Agotado": "agotado"}
                if tag != map_f.get(estado_filter, ""):
                    continue

            self.tree.insert("", "end", iid=str(row["id"]), tags=(tag,), values=(
                row["id"], row["fabricante"], row["referencia"],
                row["descripcion"] or "", qty,
                f"${row['valor_unitario']:,.0f}",
                f"${row['valor_total']:,.0f}",
                row["ubicacion"] or "", row["maquina"] or "", estado
            ))
            count += 1

        self.lbl_count.config(text=f"Mostrando {count} producto(s)")

    def _on_select(self, event):
        sel = self.tree.selection()
        if not sel:
            return
        item_id = int(sel[0])
        row = self.db.get_item_by_id(item_id)
        if not row:
            return

        self.detail_text.config(state="normal")
        self.detail_text.delete("1.0", "end")
        self.detail_text.insert("end",
            f"Fabricante:\n{row['fabricante']}\n\n"
            f"Referencia:\n{row['referencia']}\n\n"
            f"Descripción:\n{row['descripcion'] or '—'}\n\n"
            f"Cantidad:\n{row['cantidad']} uds\n\n"
            f"Valor Unitario:\n${row['valor_unitario']:,.0f}\n\n"
            f"Valor Total:\n${row['valor_total']:,.0f}\n\n"
            f"Ubicación:\n{row['ubicacion'] or '—'}\n\n"
            f"Máquina:\n{row['maquina'] or '—'}\n\n"
            f"Stock Mínimo:\n{row['stock_minimo']}\n"
            f"Stock Crítico:\n{row['stock_critico']}\n\n"
            f"Observaciones:\n{row['observaciones'] or '—'}"
        )
        self.detail_text.config(state="disabled")

        # Mostrar imagen si existe
        if row["imagen_path"] and os.path.exists(row["imagen_path"]):
            try:
                from PIL import Image, ImageTk
                img = Image.open(row["imagen_path"])
                img.thumbnail((180, 140))
                photo = ImageTk.PhotoImage(img)
                self.img_label.config(image=photo, text="")
                self.img_label.image = photo
            except ImportError:
                self.img_label.config(text="📦", font=("Arial", 40), image="")
        else:
            self.img_label.config(text="📦", font=("Arial", 40), image="")

        self.selected_id = item_id

    def _on_double_click(self, event):
        if self.selected_id:
            self._edit_item(self.selected_id)

    def _edit_item(self, item_id):
        row = self.db.get_item_by_id(item_id)
        if not row:
            return
        win = tk.Toplevel(self.parent)
        win.title(f"Editar: {row['referencia']}")
        win.configure(bg=BG_DARK)
        win.geometry("400x480")
        win.grab_set()

        frame = tk.Frame(win, bg=BG_CARD, padx=20, pady=20)
        frame.pack(fill="both", expand=True, padx=16, pady=16)

        tk.Label(frame, text="✏ EDITAR PRODUCTO", font=("Arial", 11, "bold"),
                 bg=BG_CARD, fg=ACCENT).grid(row=0, columnspan=2, sticky="w", pady=(0,12))

        fields = [
            ("Fabricante", "fabricante"),
            ("Referencia", "referencia"),
            ("Descripción", "descripcion"),
            ("Valor Unitario", "valor_unitario"),
            ("Stock Mínimo", "stock_minimo"),
            ("Stock Crítico", "stock_critico"),
            ("Ubicación", "ubicacion"),
            ("Máquina", "maquina"),
            ("Observaciones", "observaciones"),
        ]
        entries = {}
        for i, (lbl, key) in enumerate(fields, 1):
            tk.Label(frame, text=lbl, font=("Arial", 9), bg=BG_CARD, fg=TEXT_DIM).grid(
                row=i, column=0, sticky="w", padx=(0,8), pady=2)
            e = tk.Entry(frame, font=("Arial", 10), bg=BG_INPUT, fg=TEXT,
                         insertbackground=TEXT, relief="flat", bd=6, width=26)
            e.grid(row=i, column=1, sticky="ew", pady=2)
            val = row[key]
            if val is not None:
                e.insert(0, str(val))
            entries[key] = e

        def save():
            data = {k: e.get().strip() for k, e in entries.items()}
            try:
                data["valor_unitario"] = float(data["valor_unitario"]) if data["valor_unitario"] else 0
                data["stock_minimo"] = float(data["stock_minimo"]) if data["stock_minimo"] else 2
                data["stock_critico"] = float(data["stock_critico"]) if data["stock_critico"] else 1
                data["imagen_path"] = row["imagen_path"] or ""
            except ValueError:
                messagebox.showerror("Error", "Valores numéricos inválidos")
                return
            self.db.update_item(item_id, data)
            win.destroy()
            self._load_table()

        tk.Button(frame, text="💾 Guardar Cambios", font=("Arial", 10, "bold"),
                  bg=SUCCESS, fg=BG_DARK, relief="flat", cursor="hand2",
                  command=save, pady=8).grid(row=len(fields)+1, columnspan=2, pady=12)

    def _quick_add(self):
        from modules.entradas import EntradasModule
        win = tk.Toplevel(self.parent)
        win.title("Nueva Entrada")
        win.configure(bg=BG_DARK)
        win.geometry("700x600")
        win.grab_set()
        EntradasModule(win, self.db, self.user)

    def _export_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showerror("Error", "openpyxl no instalado.")
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="inventario_export.xlsx"
        )
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventario"

        headers = ["ID", "Fabricante", "Referencia", "Descripción", "Cantidad",
                   "Valor Unitario", "Valor Total", "Ubicación", "Máquina", "Estado"]
        header_fill = PatternFill("solid", start_color="0D1117")
        bold_font = Font(bold=True, color="00D4FF")

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = bold_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        items = self.db.get_inventario()
        for r_idx, row in enumerate(items, 2):
            qty = row["cantidad"]
            if qty == 0: estado = "AGOTADO"
            elif qty <= row["stock_critico"]: estado = "CRÍTICO"
            elif qty <= row["stock_minimo"]: estado = "BAJO"
            else: estado = "Normal"

            ws.append([
                row["id"], row["fabricante"], row["referencia"],
                row["descripcion"], qty, row["valor_unitario"],
                row["valor_total"], row["ubicacion"], row["maquina"], estado
            ])

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        wb.save(path)
        messagebox.showinfo("Exportado", f"Inventario exportado:\n{path}")

    def _sort(self, col):
        # Ordenamiento básico por columna
        data = [(self.tree.set(child, col), child) for child in self.tree.get_children("")]
        try:
            data.sort(key=lambda x: float(x[0].replace("$","").replace(",","")) if x[0].replace("$","").replace(",","").replace(".","").isdigit() else x[0].lower())
        except:
            data.sort(key=lambda x: x[0].lower())
        for idx, (_, child) in enumerate(data):
            self.tree.move(child, "", idx)
